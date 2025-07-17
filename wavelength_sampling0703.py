import numpy as np
import matplotlib.pyplot as plt
from astropy.io import fits
from spectres import spectres
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# 픽셀 -> 파장 변환식
def pixel_to_wavelength(pixels):
    return 3.910229e3 + 8.855374e-1 * pixels + 1.706681e-6 * pixels ** 2

# FITS에서 1D 스펙트럼 데이터 로드
def load_spectrum_from_fits(file_path):
    with fits.open(file_path) as hdul:
        data = hdul[0].data
    fluxes = data
    pixels = np.arange(len(data))
    return pixels, fluxes

# 보간 수행 및 전후 비교 플롯
def resample_with_spectres(file_path, wl_step, show=True):
    pixels, fluxes = load_spectrum_from_fits(file_path)
    wavelengths = pixel_to_wavelength(pixels)

    # 균일 파장 생성
    wl_start, wl_end = wavelengths[0], wavelengths[-1]
    uniform_wavelengths = np.arange(wl_start, wl_end, wl_step)

    # 플럭스 밀도로 변환 (적분값 보존)
    delta_lambda = np.gradient(wavelengths)
    flux_density = fluxes / delta_lambda

    # spectres 사용
    resampled_flux_density = spectres(uniform_wavelengths, wavelengths, flux_density)
    resampled_delta_lambda = np.gradient(uniform_wavelengths)
    resampled_fluxes = resampled_flux_density * resampled_delta_lambda

    if show:
        # 기존 그래프: Flux vs. Pixel Index
        fig, axs = plt.subplots(1, 2, figsize=(14, 5))
        axs[0].plot(pixels, fluxes, label='Original', color='royalblue')
        axs[0].set_title('Original Spectrum (Pixel Index)')
        axs[0].set_xlabel('Pixel Index')
        axs[0].set_ylabel('Flux')
        axs[0].grid(True)

        axs[1].plot(uniform_wavelengths, resampled_fluxes, label='Resampled', color='crimson')
        axs[1].set_title('Resampled Spectrum (Spectres)')
        axs[1].set_xlabel('Wavelength (Å)')
        axs[1].set_ylabel('Flux')
        axs[1].grid(True)

        plt.tight_layout()
        plt.show()

        # 겹쳐진 Flux vs. Wavelength 그래프
        plt.figure(figsize=(8, 5))
        plt.plot(wavelengths, fluxes, label="Original Flux", color="royalblue", linestyle='--', marker='o', markersize=6)
        plt.plot(uniform_wavelengths, resampled_fluxes, label="Resampled Flux", color="crimson", linestyle='--', marker='o', markersize=4)
        plt.xlabel("Wavelength (Å)")
        plt.ylabel("Flux")
        plt.title("Original vs. Resampled Spectrum (Flux)")
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.show()

        # 👉 추가된 그래프: Flux Density vs. Wavelength
        plt.figure(figsize=(8, 5))
        plt.plot(wavelengths, flux_density, label="Original Flux Density", color="navy", linestyle='-', marker='x', markersize=4)
        plt.plot(uniform_wavelengths, resampled_flux_density, label="Resampled Flux Density", color="darkred", linestyle='-', marker='x', markersize=4)
        plt.xlabel("Wavelength (Å)")
        plt.ylabel("Flux Density")
        plt.title("Original vs. Resampled Flux Density")
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.show()

    return pixels, fluxes, wavelengths, flux_density, uniform_wavelengths, resampled_fluxes, resampled_flux_density

# 엑셀 저장
def save_flux_data_to_excel(pixel_indices, original_flux, pixel_wavelengths, original_flux_density,
                            resampled_wavelengths, resampled_flux, resampled_flux_density,
                            output_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Original Spectrum"

    df1 = pd.DataFrame({
        "Pixel Index": pixel_indices,
        "Original Flux": original_flux,
        "Wavelength (Å)": pixel_wavelengths,
        "Flux Density": original_flux_density
    })
    for r in dataframe_to_rows(df1, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet("Resampled Spectrum")
    df2 = pd.DataFrame({
        "Wavelength (Å)": resampled_wavelengths,
        "Resampled Flux": resampled_flux,
        "Flux Density": resampled_flux_density
    })
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws2.append(r)

    # 적분값 비교용 시트
    ws3 = wb.create_sheet("Integration Checker")
    ws3['A1'] = "λ Start"
    ws3['B1'] = "λ End"
    ws3['C1'] = "Original Flux Sum"
    ws3['D1'] = "Resampled Flux Sum"
    ws3['A2'] = 3920
    ws3['B2'] = 4000

    n1 = len(df1) + 1
    n2 = len(df2) + 1

    ws3['C2'] = f"""=SUMIFS('Original Spectrum'!B2:B{n1}, 'Original Spectrum'!C2:C{n1}, ">="&A2, 'Original Spectrum'!C2:C{n1}, "<="&B2)"""
    ws3['D2'] = f"""=SUMIFS('Resampled Spectrum'!B2:B{n2}, 'Resampled Spectrum'!A2:A{n2}, ">="&A2, 'Resampled Spectrum'!A2:A{n2}, "<="&B2)"""

    wb.save(output_path)

# 실행
file_path = r"D:\________작업중_TCrB_코드\20250223_1\spec20250223_1\extrac_geo_crr_prepro_crop_Arc-001.fit_geo_crr_prepro_crop_PiSer-001.fit.fits"
output_path = r"D:\________작업중_TCrB_코드\tcrb 20250701\spectrum_analysis0703.xlsx"

pixels, fluxes, wavelengths_pre, flux_density_pre, wavelengths_post, flux_post, flux_density_post = resample_with_spectres(
    file_path=file_path,
    wl_step=0.45,
    show=True
)

save_flux_data_to_excel(
    pixel_indices=pixels,
    original_flux=fluxes,
    pixel_wavelengths=wavelengths_pre,
    original_flux_density=flux_density_pre,
    resampled_wavelengths=wavelengths_post,
    resampled_flux=flux_post,
    resampled_flux_density=flux_density_post,
    output_path=output_path
)
